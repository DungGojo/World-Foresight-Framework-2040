#!/usr/bin/env python3
"""
Build-time data pipeline: Final Data.xlsx  ->  src/data/*.json

Emits, for all five topics (no runtime XLSX parsing, no backend):

  1. timeseries.json        — the Level-3 explorer feed, shared by every topic.
     Nested proxy -> market -> scenario -> [[year, value, lower_ci, upper_ci], ...]
     (nesting instead of tidy rows keeps it small — no repeated keys).
     Plus proxy metadata, per-country statistical summaries and a market
     catalog (code -> country name) for the pickers. Each proxy carries the
     list of topics it belongs to, read from the 'Topics' column.

  2. {topic}-figures.json   — the exact series for each topic's Level-2 charts,
     one file per topic (power / tech / planet / people / economy), plus
     cross-figures.json for the hub's cross-topic compound-risk section.

     Three classes of figure value, tagged so the site can tell them apart:
       "computed"  — recomputed here from the sheet, mirroring the notebook's
                     Analysis_Functions (share_of_world, quadrant_map,
                     rank_market_relevance). Anchor checks print on every run.
       "spec"      — editorial numbers transcribed from the topic docx, where
                     the finding needs a judgment call the sheet cannot make.
       "external"  — sourced from IPCC/IMF/ITU/etc, never from the pipeline.
                     Per MODEL_LIMITATIONS.md, absolute 2040 levels must come
                     from an external authority, not from our damped forecast.

Run:  python3 scripts/build_data.py
Uses openpyxl only (the repo's base pandas has a numpy2 conflict).
"""
import json, math, os, pickle, re, sys
from collections import defaultdict

# --figures-only skips the Level-3 timeseries + catalog files and emits just the
# {topic}-figures.json set. Those timeseries depend only on the spreadsheet, so
# when the edit was to topic_figures.py (a chart's note/annotation//label) there
# is nothing in them to change — and skipping ~12 MB of JSON serialisation is
# most of the runtime. Used by the dev-server watcher in vite.config.js.
FIGURES_ONLY = "--figures-only" in sys.argv

HERE = os.path.dirname(os.path.abspath(__file__))
APP  = os.path.dirname(HERE)
XLSX = os.path.normpath(os.path.join(APP, "..", "..", "Final Data.xlsx"))
OUT  = os.path.join(APP, "src", "data")
os.makedirs(OUT, exist_ok=True)

sys.path.insert(0, HERE)

MARKET_NAMES = {
    "ARE":"United Arab Emirates","ARG":"Argentina","AUS":"Australia","BGD":"Bangladesh",
    "BRA":"Brazil","CAN":"Canada","CHN":"China","COD":"DR Congo","DEU":"Germany",
    "EGY":"Egypt","ETH":"Ethiopia","FRA":"France","GBR":"United Kingdom","GLO":"World",
    "IDN":"Indonesia","IND":"India","IRN":"Iran","ISR":"Israel","ITA":"Italy","JPN":"Japan",
    "KAZ":"Kazakhstan","KEN":"Kenya","KOR":"South Korea","MEX":"Mexico","NGA":"Nigeria",
    "NLD":"Netherlands","PAK":"Pakistan","POL":"Poland","RUS":"Russia","SAU":"Saudi Arabia",
    "TUR":"Turkey","UKR":"Ukraine","USA":"United States","VNM":"Vietnam","ZAF":"South Africa",
}

# 'Topics' column holds bracketed numeric lists ('[1]', '[1, 2]'); 1-5 map to:
TOPIC_BY_NUMBER = {1: "power", 2: "tech", 3: "planet", 4: "people", 5: "economy"}
TOPIC_IDS = ["power", "tech", "planet", "people", "economy"]

def rnd(v, n=4):
    if v is None: return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f != f:  # NaN
        return None
    r = round(f, n)
    return int(r) if r == int(r) else r

def die(msg):
    print("ERROR:", msg); sys.exit(1)

def parse_topics(cell):
    """'[1, 2]' -> ['power','tech'].  Unparseable/empty -> []."""
    if cell is None:
        return []
    nums = re.findall(r"\d+", str(cell))
    out = []
    for n in nums:
        slug = TOPIC_BY_NUMBER.get(int(n))
        if slug and slug not in out:
            out.append(slug)
    return out

# ---------------------------------------------------------------- load sheet
try:
    import openpyxl
except ImportError:
    die("openpyxl not installed (pip install openpyxl)")
if not os.path.exists(XLSX):
    die(f"Final Data.xlsx not found at {XLSX}")

# ---------------------------------------------------------------- parse cache
# Parsing the 380k-row sheet is ~20s and dominates the run; the parsed result
# only changes when the spreadsheet does. Cache it, keyed on the file's exact
# size + mtime, so repeated runs (especially the dev-server watcher rebuilding
# after a topic_figures.py edit) skip straight to the analysis in well under a
# second. Any change to Final Data.xlsx invalidates the key automatically;
# --no-cache forces a full re-parse.
CACHE_PATH = os.path.join(HERE, ".cache", "parsed.pkl")
CACHE_VERSION = 1   # bump when the cached structures below change shape

def _xlsx_stamp():
    st = os.stat(XLSX)
    return {"v": CACHE_VERSION, "mtime_ns": st.st_mtime_ns, "size": st.st_size}

_cached = None
if "--no-cache" not in sys.argv and os.path.exists(CACHE_PATH):
    try:
        with open(CACHE_PATH, "rb") as f:
            _blob = pickle.load(f)
        if _blob.get("stamp") == _xlsx_stamp():
            _cached = _blob
    except Exception:
        _cached = None   # unreadable/stale cache is never fatal - just re-parse

if _cached is not None:
    print(f"reading cached parse of {os.path.basename(XLSX)} "
          f"(pass --no-cache to force a re-read)")
    proxy_meta        = _cached["proxy_meta"]
    nested            = _cached["nested"]
    raw               = _cached["raw"]
    proxy_unit        = _cached["proxy_unit"]
    proxy_ids_present = _cached["proxy_ids_present"]
    statistics        = _cached["statistics"]
    print(f"  {len(raw)} rows \u00b7 {len(proxy_ids_present)} base proxies \u00b7 {len(nested)} series groups (cached)")
else:
    print(f"reading {XLSX}")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # proxy catalog from 'Proxy Information' (this is where 'Topics' lives)
    proxy_meta = {}
    if "Proxy Information" in wb.sheetnames:
        ws = wb["Proxy Information"]
        it = ws.iter_rows(values_only=True)
        header = next(it)
        hi = {h: i for i, h in enumerate(header)}
        if "Topics" not in hi:
            die("column 'Topics' missing from Proxy Information — topic tagging would "
                "silently fall back to power for every proxy")
        for r in it:
            if not r or r[hi.get("id")] is None:
                continue
            pid = str(r[hi["id"]]).strip()
            proxy_meta[pid] = {
                "id": pid,
                "name": r[hi.get("proxy_name")] if "proxy_name" in hi else pid,
                "description": r[hi.get("proxy_description")] if "proxy_description" in hi else "",
                "source": r[hi.get("data_source")] if "data_source" in hi else "",
                "topics": parse_topics(r[hi["Topics"]]),
                "note": r[hi.get("NOTE")] if "NOTE" in hi else "",
            }


    # main data
    ws = wb["Final Full Data"]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    hi = {h: i for i, h in enumerate(header)}
    for req in ("proxy_id","id","market","year","value","labels","scenario","lower_ci","upper_ci"):
        if req not in hi:
            die(f"column '{req}' missing from Final Full Data")

    nested = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    proxy_unit = {}
    proxy_ids_present = set()
    rows = 0
    raw = {}   # raw[(base_id, market, year, scenario)] = value

    for r in it:
        if r is None or r[hi["proxy_id"]] is None:
            continue
        base  = str(r[hi["id"]]).strip()
        mkt   = str(r[hi["market"]]).strip()
        yr    = r[hi["year"]]
        val   = r[hi["value"]]
        unit  = r[hi["labels"]]
        scen  = str(r[hi["scenario"]]).strip()
        lo    = r[hi["lower_ci"]]
        hh    = r[hi["upper_ci"]]
        if yr is None or val is None:
            continue
        yr = int(yr)
        nested[base][mkt][scen].append([yr, rnd(val), rnd(lo), rnd(hh)])
        if base not in proxy_unit and unit:
            proxy_unit[base] = str(unit)
        proxy_ids_present.add(base)
        raw[(base, mkt, yr, scen)] = float(val)
        rows += 1

    print(f"  {rows} rows · {len(proxy_ids_present)} base proxies · {len(nested)} series groups")

    for base in nested:
        for mkt in nested[base]:
            for scen in nested[base][mkt]:
                nested[base][mkt][scen].sort(key=lambda x: x[0])

    STAT_FIELDS = ("change_direction", "change_speed", "uncertainty_level", "cagr_p1y", "cagr_p3y", "cagr_full")
    statistics = defaultdict(dict)
    if "Statistical Data" in wb.sheetnames:
        ws_stats = wb["Statistical Data"]
        stat_rows = ws_stats.iter_rows(values_only=True)
        stat_header = next(stat_rows)
        si = {h: i for i, h in enumerate(stat_header)}
        if all(key in si for key in ("id", "market", *STAT_FIELDS)):
            for r in stat_rows:
                if not r or r[si["id"]] in (None, "id") or r[si["market"]] in (None, "market"):
                    continue
                base = str(r[si["id"]]).strip()
                market = str(r[si["market"]]).strip()
                statistics[base][market] = {
                    field: (rnd(r[si[field]]) if field.startswith("cagr_") else r[si[field]])
                    for field in STAT_FIELDS
                }

    # Store plain dicts: `nested` is built from defaultdicts whose lambda
    # factories are not picklable, and everything downstream only ever reads.
    os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
    try:
        with open(CACHE_PATH, "wb") as f:
            pickle.dump({
                "stamp": _xlsx_stamp(),
                "proxy_meta": dict(proxy_meta),
                "nested": {b: {m: dict(sc) for m, sc in mk.items()} for b, mk in nested.items()},
                "raw": raw,
                "proxy_unit": dict(proxy_unit),
                "proxy_ids_present": set(proxy_ids_present),
                "statistics": {b: dict(v) for b, v in statistics.items()},
            }, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception as e:
        print(f"  (could not write parse cache: {e})")

tally = defaultdict(int)
for m in proxy_meta.values():
    for t in m["topics"]:
        tally[t] += 1
untagged = [p for p, m in proxy_meta.items() if not m["topics"]]
print(f"  proxy topics: " + "  ".join(f"{t}={tally[t]}" for t in TOPIC_IDS) +
      f"  (untagged={len(untagged)})")
if untagged:
    print(f"    untagged ids: {', '.join(sorted(untagged)[:12])}")

# ---------------------------------------------------------------- proxy catalog
def group_of(base):
    return base.split("_")[0]


def humanize_catalog_copy(text, *, is_name=False):
    """Remove editorial dash separators from user-facing proxy metadata."""
    if not isinstance(text, str):
        return text
    return text.replace(" — ", ": " if is_name else ", ")


catalog = []
for base in sorted(proxy_ids_present, key=lambda b: (b.split("_")[0], b)):
    meta = proxy_meta.get(group_of(base), proxy_meta.get(base, {}))
    topics = meta.get("topics", [])
    catalog.append({
        "id": base,
        "group": group_of(base),
        "name": humanize_catalog_copy(meta.get("name") or base, is_name=True),
        # `topic` kept for backwards compatibility; `topics` is authoritative
        # because a proxy can belong to more than one topic (e.g. '[1, 2]').
        "topic": topics[0] if topics else "",
        "topics": topics,
        "unit": proxy_unit.get(base, ""),
        "description": humanize_catalog_copy(meta.get("description", "")),
        "source": meta.get("source", ""),
    })

markets_present = sorted({m for base in nested for m in nested[base] if m != "GLO"})
has_glo = any("GLO" in nested[base] for base in nested)
market_catalog = ([{"code": "GLO", "name": "World"}] if has_glo else []) + \
    [{"code": m, "name": MARKET_NAMES.get(m, m)} for m in markets_present]

# One timeseries file per topic. The full set is ~12 MB, far too big to pull
# into the browser at once; split by topic each file is 1.6-2.7 MB and only the
# topic being explored is fetched. A proxy tagged '[1, 2]' appears in both files.
SCENARIOS = ["historical", "main_scenario", "optimistic_scenario", "pessimistic_scenario"]
for topic in ([] if FIGURES_ONLY else TOPIC_IDS):
    ids = [c["id"] for c in catalog if topic in c["topics"]]
    idset = set(ids)
    topic_markets = sorted({m for b in idset if b in nested for m in nested[b] if m != "GLO"})
    topic_has_glo = any("GLO" in nested[b] for b in idset if b in nested)
    payload = {
        "meta": {"topic": topic, "years": [2000, 2040], "scenarios": SCENARIOS},
        "proxies": [c for c in catalog if c["id"] in idset],
        "markets": ([{"code": "GLO", "name": "World"}] if topic_has_glo else []) +
                   [{"code": m, "name": MARKET_NAMES.get(m, m)} for m in topic_markets],
        "statistics": {b: dict(statistics[b]) for b in statistics if b in idset},
        "series": {b: {m: dict(nested[b][m]) for m in nested[b]} for b in idset if b in nested},
    }
    path = os.path.join(OUT, f"{topic}-timeseries.json")
    with open(path, "w") as f:
        json.dump(payload, f, separators=(",", ":"))
    print(f"  wrote {topic}-timeseries.json ({os.path.getsize(path)/1024:.0f} KB, {len(ids)} proxies)")

# A tiny index of every proxy across all five topics, so the data explorer can
# search all 190 indicators without pulling any of the ~2.5 MB series files.
# `topic` names the file that actually holds the series, fetched on selection.
if not FIGURES_ONLY:
    catalog_payload = {
        "proxies": [{k: c[k] for k in ("id", "name", "unit", "description", "source", "topic", "topics")}
                    for c in catalog if c["topics"]],
        "markets": market_catalog,
    }
    cat_path = os.path.join(OUT, "catalog.json")
    with open(cat_path, "w") as f:
        json.dump(catalog_payload, f, separators=(",", ":"))
    print(f"  wrote catalog.json ({os.path.getsize(cat_path)/1024:.0f} KB, "
          f"{len(catalog_payload['proxies'])} proxies)")

# ================================================================ analysis
COUNTRY_MARKETS = list(markets_present)          # GLO excluded from all math
SCEN_CHAIN = ("historical", "main_scenario")     # prefer actuals at the seam

def slice_values(pid, year, scen_chain=SCEN_CHAIN):
    """{market: value} for one proxy at one year, actuals preferred."""
    out = {}
    for m in COUNTRY_MARKETS:
        for s in scen_chain:
            v = raw.get((pid, m, year, s))
            if v is not None:
                out[m] = v
                break
    return out

def global_at(pid, year, scen_chain=SCEN_CHAIN):
    """The GLO ('World') row for one proxy at one year, or None.

    slice_values deliberately drops GLO so it can never contaminate a share
    calculation, but the treaty-ratification and peacekeeping series are only
    published as world totals, so they need their own reader."""
    for s in scen_chain:
        v = raw.get((pid, "GLO", year, s))
        if v is not None:
            return v
    return None

def years_for(pid):
    return sorted({k[2] for k in raw if k[0] == pid})

# ---- share of world (mirrors Analysis_Functions/share_of_world.share_trajectory)
def proxy_share_by_year(base):
    """{year: {market: share_pct}} — value / sum across the country markets."""
    out = defaultdict(dict)
    for y in years_for(base):
        vals = slice_values(base, y)
        total = sum(vals.values())
        if total > 0:
            for m, v in vals.items():
                out[y][m] = 100.0 * v / total
    return out

def composite_share(proxies):
    """Mean of per-proxy shares across a topic's additive proxy set."""
    per = {p: proxy_share_by_year(p) for p in proxies}
    years = sorted(set().union(*[set(per[p].keys()) for p in proxies]))
    comp = {}
    for y in years:
        comp[y] = {}
        for m in COUNTRY_MARKETS:
            shares = [per[p][y][m] for p in proxies if y in per[p] and m in per[p][y]]
            if shares:
                comp[y][m] = sum(shares) / len(shares)
    return comp

def direct_share(base):
    """For proxies that ARE already a share of world (D143/D151/D152/D159).
    These must be read through, never re-normalised against a 34-country sum."""
    out = {}
    for y in years_for(base):
        vals = slice_values(base, y)
        if vals:
            out[y] = dict(vals)
    return out

def lines_from(comp, markets, decimals=2):
    return [{"market": m, "name": MARKET_NAMES.get(m, m),
             "values": [[y, round(comp[y][m], decimals)] for y in sorted(comp) if m in comp[y]]}
            for m in markets]

def anchors_from(comp, market, years):
    return [round(comp.get(y, {}).get(market, float("nan")), 1) for y in years]

def check(label, got, want, tol=0.15):
    """Print a computed-vs-docx anchor check so drift is visible on every run."""
    ok = all(g is not None and w is not None and abs(g - w) <= tol for g, w in zip(got, want))
    print(f"    {'OK ' if ok else '!! '}{label}: computed {got}  docx {want}")
    if not ok:
        DRIFT.append(label)

DRIFT = []

# ---- composites (mirrors rank_market_relevance.standardize_to_100)
def standardize_to_100(vals):
    valid = [v for v in vals if v is not None]
    n = len(valid)
    if n == 0:
        return [None] * len(vals)
    if n < 2:
        return [50.0 if v is not None else None for v in vals]
    mean = sum(valid) / n
    std = math.sqrt(sum((v - mean) ** 2 for v in valid) / (n - 1))
    if std < 1e-9:
        return [50.0 if v is not None else None for v in vals]
    return [None if v is None else max(0.0, min(100.0, ((v - mean) / std + 2) / 4 * 100))
            for v in vals]

def composite_index(config, year):
    """config = [(proxy_id, 'positive'|'negative')] -> {market: 0-100 score}.
    Equal-weighted z-scores per proxy, averaged, then re-standardised —
    identical to Analysis_Functions/market_relevance.rank_market_relevance."""
    per_proxy = {}
    for pid, direction in config:
        vals = slice_values(pid, year)
        if not vals:
            print(f"    !! composite: {pid} has no data at {year}")
            continue
        col = [vals.get(m) for m in COUNTRY_MARKETS]
        if direction == "negative":
            col = [None if v is None else -v for v in col]
        per_proxy[pid] = standardize_to_100(col)
    if not per_proxy:
        return {}
    agg = []
    for i in range(len(COUNTRY_MARKETS)):
        got = [per_proxy[p][i] for p in per_proxy if per_proxy[p][i] is not None]
        agg.append(sum(got) / len(got) if got else None)
    final = standardize_to_100(agg)
    return {m: final[i] for i, m in enumerate(COUNTRY_MARKETS) if final[i] is not None}

def correlation(a, b, keys):
    n = len(keys)
    ma = sum(a[k] for k in keys) / n
    mb = sum(b[k] for k in keys) / n
    num = sum((a[k] - ma) * (b[k] - mb) for k in keys)
    den = math.sqrt(sum((a[k] - ma) ** 2 for k in keys) * sum((b[k] - mb) ** 2 for k in keys))
    return num / den if den else 0.0

def quadrant_map(x_config, y_config, year, x_label, y_label, names, split=50.0):
    """Mirrors Analysis_Functions/quadrant.quadrant_map + alignment_score."""
    X = composite_index(x_config, year)
    Y = composite_index(y_config, year)
    keys = sorted(set(X) & set(Y))
    points = []
    counts = defaultdict(int)
    for m in keys:
        xh, yh = X[m] >= split, Y[m] >= split
        q = names.get((xh, yh), "")
        counts[q] += 1
        points.append({
            "market": m, "name": MARKET_NAMES.get(m, m),
            "x": round(X[m], 1), "y": round(Y[m], 1),
            "quadrant": q, "gap": round(X[m] - Y[m], 1),
            # Appendix B: membership near the split is sensitive, so flag it
            # rather than presenting a hard assignment.
            "borderline": abs(X[m] - split) < 5 or abs(Y[m] - split) < 5,
        })
    points.sort(key=lambda p: -p["gap"])
    return {
        "xLabel": x_label, "yLabel": y_label, "split": split, "year": year,
        "points": points, "counts": dict(counts),
        "correlation": round(correlation(X, Y, keys), 3),
        "quadrants": {f"{int(k[0])}{int(k[1])}": v for k, v in names.items()},
        "computed": True,
    }

def members(qmap, quadrant):
    return sorted(p["market"] for p in qmap["points"] if p["quadrant"] == quadrant)

# ---- nearest-anchor clustering (mirrors Analysis_Functions/cluster_to_anchor)
CLUSTER_ANCHORS = {"USA": ("D8_1", "D9_1"), "CHN": ("D8_2", "D9_2"),
                   "RUS": ("D8_3", "D9_3"), "IND": ("D8_4", "D9_4")}

def _zscore(vals):
    """{market: value} -> {market: z}, sample sd, matching pandas std(ddof=1)."""
    xs = list(vals.values())
    n = len(xs)
    if n < 2:
        return {m: 0.0 for m in vals}
    mean = sum(xs) / n
    sd = math.sqrt(sum((x - mean) ** 2 for x in xs) / (n - 1))
    return {m: (v - mean) / (sd + 1e-9) for m, v in vals.items()}

def cluster_to_anchor(year, anchors=CLUSTER_ANCHORS):
    """Assign each non-anchor market to the anchor power it sits closest to.

    Mirrors Analysis_Functions/cluster_to_anchor.cluster_to_anchor: z-score
    UN-vote agreement (D8_*) and trade share (D9_*) across markets so the two
    units are comparable, average the pair into a proximity per anchor, drop the
    self-anchor, then take the argmax. `margin` (top1 - top2) is how confident
    the assignment is."""
    z = {pid: _zscore(slice_values(pid, year))
         for pair in anchors.values() for pid in pair}
    out = {}
    for m in COUNTRY_MARKETS:
        if m in anchors:
            continue
        prox = {}
        for label, ids in anchors.items():
            parts = [z[pid][m] for pid in ids if m in z[pid]]
            if parts:
                prox[label] = sum(parts) / len(parts)
        if not prox:
            continue
        ranked = sorted(prox.items(), key=lambda kv: -kv[1])
        out[m] = {
            "market": m, "name": MARKET_NAMES.get(m, m), "cluster": ranked[0][0],
            "proximity": round(ranked[0][1], 3),
            "margin": round(ranked[0][1] - ranked[1][1], 3) if len(ranked) > 1 else None,
        }
    return out

def cluster_members(year, anchor):
    """[{market, name, proximity, margin}] for one pole, closest first."""
    rows = [r for r in cluster_to_anchor(year).values() if r["cluster"] == anchor]
    rows.sort(key=lambda r: -r["proximity"])
    return [{k: r[k] for k in ("market", "name", "proximity", "margin")} for r in rows]


# ---- alignment tilt (mirrors Analysis_Functions/cluster_to_anchor.alignment_tilt)
TILT_DOMAINS = {"votes": ("D8_1", "D8_2"), "trade": ("D9_1", "D9_2")}
ALIGN_THRESHOLD, HEDGE_THRESHOLD = 0.35, 0.60

def alignment_tilt(year, domains=TILT_DOMAINS, exclude=("USA", "CHN")):
    """{market: {tilt, balance, posture, split, votes, trade}} for one year.

    Per domain: tilt = (west - east) / (west + east), so +1 is fully US-facing
    and -1 fully China-facing; balance = 2*min/(west+east), 0 one-sided and 1
    evenly split. The composite is the mean across domains, and `split` flags
    the countries whose domains point opposite ways — literally aligned one way
    on votes and the other way on money."""
    vals = {pid: slice_values(pid, year) for pair in domains.values() for pid in pair}
    out = {}
    for m in COUNTRY_MARKETS:
        if m in exclude:
            continue
        per, balances = {}, []
        for name, (west_id, east_id) in domains.items():
            if m not in vals[west_id] or m not in vals[east_id]:
                continue
            west, east = vals[west_id][m], vals[east_id][m]
            total = west + east
            if abs(total) < 1e-9:
                continue
            per[name] = (west - east) / total
            balances.append(2 * min(west, east) / total)
        if not per:
            continue
        tilt = sum(per.values()) / len(per)
        balance = sum(balances) / len(balances)
        if tilt >= ALIGN_THRESHOLD:
            posture = "west"
        elif tilt <= -ALIGN_THRESHOLD:
            posture = "east"
        elif balance >= HEDGE_THRESHOLD:
            posture = "hedge"
        else:
            posture = "auto"
        signs = {(v > 0) - (v < 0) for v in per.values()}
        out[m] = {
            "market": m, "name": MARKET_NAMES.get(m, m),
            "tilt": round(tilt, 3), "balance": round(balance, 3), "posture": posture,
            "split": len(signs) > 1,
            **{k: round(v, 3) for k, v in per.items()},
        }
    return out

def posture_counts(year):
    rows = alignment_tilt(year)
    counts = defaultdict(int)
    for r in rows.values():
        counts[r["posture"]] += 1
    return dict(counts), rows

def mean_tilt(year):
    rows = alignment_tilt(year)
    return round(sum(r["tilt"] for r in rows.values()) / len(rows), 3) if rows else None

# ---- per-domain world shares, for the power-signature radar
RADAR_AXES = [
    ("Military", "D1", "share"),
    ("Technology", "D16", "share"),
    ("Industrial capacity", "D159", "direct"),
    ("Trade", "D2", "share"),
    ("GDP", "D4", "share"),
]

def radar_axis_values(year, axes=RADAR_AXES):
    """{axis_label: {market: % of the 34-country world}} for one year.

    'share' proxies are levels normalised here; 'direct' ones (D159) are already
    published as a share of world and must be read through, never re-normalised.
    """
    out = {}
    for label, pid, kind in axes:
        vals = slice_values(pid, year)
        if kind == "direct":
            out[label] = dict(vals)
        else:
            total = sum(vals.values())
            out[label] = {m: 100.0 * v / total for m, v in vals.items()} if total else {}
    return out
# ---------------------------------------------------------------- run analysis
print("  computed anchor checks (against the topic docx):")

# Power — composite share of world power
POWER_PROXIES = ["D1", "D2", "D4", "D5"]
power_comp = composite_share(POWER_PROXIES)
check("power share USA 2025/2030/2040", anchors_from(power_comp, "USA", [2025, 2030, 2040]), [25.2, 24.9, 24.7])
check("power share CHN 2025/2030/2040", anchors_from(power_comp, "CHN", [2025, 2030, 2040]), [19.9, 20.4, 20.7])
check("power share USA 2000/2010",      anchors_from(power_comp, "USA", [2000, 2010]),       [31.6, 28.2])
check("power share CHN 2000/2010",      anchors_from(power_comp, "CHN", [2000, 2010]),       [8.1, 14.1])

# Power — nearest-anchor clustering (the orbit map's four camps)
_clusters_2040 = cluster_to_anchor(2040)
check("orbit camps USA/IND/CHN/RUS 2040",
      [float(sum(1 for r in _clusters_2040.values() if r["cluster"] == a))
       for a in ("USA", "IND", "CHN", "RUS")], [14.0, 7.0, 6.0, 3.0], tol=0.5)

# Power — posture mix and the split-issue count (argument 3)
_post_25, _rows_25 = posture_counts(2025)
_post_40, _rows_40 = posture_counts(2040)
check("posture hedge/east/auto/west 2040",
      [float(_post_40.get(k, 0)) for k in ("hedge", "east", "auto", "west")],
      [15.0, 14.0, 2.0, 1.0], tol=0.5)
check("issue-split countries 2025/2040",
      [float(sum(1 for r in _rows_25.values() if r["split"])),
       float(sum(1 for r in _rows_40.values() if r["split"]))], [8.0, 10.0], tol=0.5)

# Technology — combined production share
TECH_PROXIES = ["D16", "D22", "D24", "D26", "D27", "D30", "D31"]
tech_comp = composite_share(TECH_PROXIES)
check("tech share CHN 2025/2040", anchors_from(tech_comp, "CHN", [2025, 2040]), [27.2, 27.5])
check("tech share USA 2025/2040", anchors_from(tech_comp, "USA", [2025, 2040]), [21.9, 21.7])

# Planet — emissions share (D62 is a level; share is over the 34-country total)
emissions_comp = composite_share(["D62"])
check("emissions share CHN 2000/2025/2040", anchors_from(emissions_comp, "CHN", [2000, 2025, 2040]), [17.6, 35.9, 38.0])
check("emissions share USA 2000/2040",      anchors_from(emissions_comp, "USA", [2000, 2040]),       [25.1, 11.9])
check("emissions share IND 2000/2040",      anchors_from(emissions_comp, "IND", [2000, 2040]),       [6.4, 10.5])

# Economy — D159/D143/D151 are already 'share of world' proxies: read through.
mfg_share = direct_share("D159")
check("mfg VA share CHN 2025/2040", anchors_from(mfg_share, "CHN", [2025, 2040]), [27.4, 35.5])
check("mfg VA share USA 2040",      anchors_from(mfg_share, "USA", [2040]),       [14.1])
check("mfg VA share JPN 2000/2040", anchors_from(mfg_share, "JPN", [2000, 2040]), [17.9, 3.9])
ppp_share = direct_share("D143")
check("PPP GDP share CHN 2000/2025/2040", anchors_from(ppp_share, "CHN", [2000, 2025, 2040]), [6.7, 20.0, 24.1])
check("PPP GDP share USA 2000/2025/2040", anchors_from(ppp_share, "USA", [2000, 2025, 2040]), [20.4, 14.7, 14.0])

# ---- the three quadrant tests (Appendix B: the three axes share no proxies)
EXPOSURE = [("D66","positive"), ("D68","positive"), ("D70","positive"),
            ("D73","positive"), ("D74","negative")]
CAPACITY = [("D92","negative"), ("D93","positive"), ("D94","negative")]
PRESSURE = [("D128","positive"), ("D123","positive"), ("D125","positive"),
            ("D131","positive"), ("D110","positive"), ("D105","positive")]
INSTITUTIONS = [("D134","positive"), ("D135","positive"), ("D136","positive"), ("D119","positive")]
STRESS  = [("D169","positive"), ("D170","positive"), ("D173","positive"),
           ("D168","positive"), ("D175","positive")]
BUFFERS = [("D171","positive"), ("D172","positive"), ("D144","positive"), ("D181","positive")]

PLANET_Q = {(True, False): "Exposed and unable", (True, True): "Exposed but capable",
            (False, True): "Sheltered and capable", (False, False): "Sheltered but unable"}
PEOPLE_Q = {(True, False): "Pressured and brittle", (True, True): "Pressured but absorbing",
            (False, True): "Calm and capable", (False, False): "Calm but brittle"}
ECON_Q   = {(True, False): "Stressed and unbuffered", (True, True): "Stressed but buffered",
            (False, True): "Steady and buffered", (False, False): "Steady but unbuffered"}

planet_q = {y: quadrant_map(EXPOSURE, CAPACITY, y, "Physical exposure", "Adaptive capacity", PLANET_Q)
            for y in (2025, 2040)}
people_q = {y: quadrant_map(PRESSURE, INSTITUTIONS, y, "Social & economic pressure", "Institutional capacity", PEOPLE_Q)
            for y in (2025, 2040)}
econ_q   = {y: quadrant_map(STRESS, BUFFERS, y, "Financial stress", "Fiscal & external buffers", ECON_Q)
            for y in (2025, 2040)}

check("planet exposure/capacity correlation 2025/2040",
      [planet_q[2025]["correlation"], planet_q[2040]["correlation"]], [-0.50, -0.50], tol=0.02)
check("people pressure/institutions correlation 2025/2040",
      [people_q[2025]["correlation"], people_q[2040]["correlation"]], [-0.49, -0.46], tol=0.02)

planet_worst = members(planet_q[2040], "Exposed and unable")
people_worst = members(people_q[2040], "Pressured and brittle")
econ_worst   = members(econ_q[2040], "Stressed and unbuffered")
check("worst-quadrant sizes planet/people/economy",
      [len(planet_worst), len(people_worst), len(econ_worst)], [8, 11, 9], tol=0)

compound = sorted(set(planet_worst) & set(people_worst) & set(econ_worst))
two_of_three = sorted(
    m for m in set(planet_worst) | set(people_worst) | set(econ_worst)
    if m not in compound
    and sum(m in s for s in (planet_worst, people_worst, econ_worst)) == 2
)
print(f"    compound risk (all three worst quadrants): {compound}  docx ['BGD','EGY','IND','KEN','PAK']")
print(f"    two of three: {two_of_three}  docx ['ETH','IRN','NGA']")
if compound != ["BGD", "EGY", "IND", "KEN", "PAK"]:
    DRIFT.append("compound-risk intersection")
subset = set(planet_worst) <= set(people_worst)
print(f"    planet worst is a strict subset of people worst: {subset}  (docx: yes)")

def hhi(pid, year):
    """Herfindahl index of one proxy's share distribution, and the implied
    'effective number of players' (10000/HHI)."""
    vals = slice_values(pid, year)
    total = sum(vals.values())
    if total <= 0:
        return None, None
    h = sum((100.0 * v / total) ** 2 for v in vals.values())
    return round(h), round(10000.0 / h, 1)

def mean_at(pid, year):
    vals = slice_values(pid, year)
    return sum(vals.values()) / len(vals) if vals else None

def series_mean(pid):
    """{year: cross-country mean} across every year the proxy covers."""
    return {y: mean_at(pid, y) for y in years_for(pid) if mean_at(pid, y) is not None}

def ranked(pid, year, top=None, reverse=True, decimals=1):
    """[{market, name, value}] sorted by value."""
    vals = slice_values(pid, year)
    rows = [{"market": m, "name": MARKET_NAMES.get(m, m), "value": round(v, decimals)}
            for m, v in vals.items()]
    rows.sort(key=lambda r: -r["value"] if reverse else r["value"])
    return rows[:top] if top else rows

def pair(pid, y0, y1, decimals=1):
    """[{market, name, from, to}] for every market present in both years."""
    a, b = slice_values(pid, y0), slice_values(pid, y1)
    return [{"market": m, "name": MARKET_NAMES.get(m, m),
             "from": round(a[m], decimals), "to": round(b[m], decimals)}
            for m in sorted(set(a) & set(b))]

computed = {
    "power_comp": power_comp, "tech_comp": tech_comp,
    "emissions_comp": emissions_comp, "mfg_share": mfg_share, "ppp_share": ppp_share,
    "planet_q": planet_q, "people_q": people_q, "econ_q": econ_q,
    "compound": compound, "two_of_three": two_of_three,
    "planet_worst": planet_worst, "people_worst": people_worst, "econ_worst": econ_worst,
    "lines_from": lines_from, "market_names": MARKET_NAMES,
    # analysis API for the figure specs
    "at": slice_values, "ranked": ranked, "pair": pair, "hhi": hhi,
    "global": global_at,
    "mean_at": mean_at, "series_mean": series_mean, "years_for": years_for,
    "share": composite_share, "direct": direct_share,
    "composite_index": composite_index, "quadrant_map": quadrant_map,
    "markets": COUNTRY_MARKETS, "check": check,
    "cluster_to_anchor": cluster_to_anchor, "cluster_members": cluster_members,
    "alignment_tilt": alignment_tilt, "posture_counts": posture_counts,
    "mean_tilt": mean_tilt, "radar_axis_values": radar_axis_values, "radar_axes": RADAR_AXES,
}

# ---------------------------------------------------------------- figures
from topic_figures import build_figures   # noqa: E402

all_figs = build_figures(computed)
for topic, figs in all_figs.items():
    path = os.path.join(OUT, f"{topic}-figures.json")
    with open(path, "w") as f:
        json.dump(figs, f, separators=(",", ":"))
    print(f"  wrote {topic}-figures.json ({os.path.getsize(path)/1024:.0f} KB, {len(figs)} figures)")

if DRIFT:
    print("\nDRIFT DETECTED — computed values no longer match the docx for:")
    for d in DRIFT:
        print(f"  - {d}")
    print("Fix the pipeline or update the docx before publishing these numbers.")
    sys.exit(1)
print("done.")
