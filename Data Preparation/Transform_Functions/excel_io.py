"""Coverage QA and the Historical Data loader.

Every proxy ends the same way: `check_coverage(df)` to see what you got, then
`load(df)` to append it to Final Data.xlsx.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from Transform_Functions.common import (
    MARKETS,
    OUTPUT_COLUMNS,
    PROJECT_ROOT,
    proxy_dim,
)

WORKBOOK = PROJECT_ROOT / "Final Data.xlsx"
SHEET = "Historical Data"
HEADERS = ["id", "proxy_id", "market", "year", "value", "labels", "metric"]


def check_coverage(df: pd.DataFrame, universe: dict | None = None,
                   verbose: bool = True) -> pd.DataFrame:
    """Report datapoints per country and which of the 34 markets are missing.

    Returns one row per country present, so the result can be inspected or
    asserted on; the printed summary is for reading at the notebook.
    """
    universe = universe or MARKETS

    # D7, D8, D9 and D11 produce several sub-proxies from one computation.
    # Reporting them together would multiply every per-country count, so each
    # sub-proxy is summarised on its own.
    if len(df):
        dims = sorted({proxy_dim(p) for p in df["proxy_id"]})
        if len(dims) > 1:
            parts = []
            for dim in dims:
                subset = df[df["proxy_id"].map(proxy_dim) == dim]
                parts.append(check_coverage(subset, universe, verbose).assign(id=dim))
                if verbose:
                    print()
            return pd.concat(parts, ignore_index=True)

    expected = set(universe)
    present = set(df["market"].unique()) if len(df) else set()

    if len(df):
        counts = df.groupby("market").size().reset_index(name="datapoints")
    else:
        counts = pd.DataFrame(columns=["market", "datapoints"])
    counts["country"] = counts["market"].map(
        lambda m: "Global series" if m == "GLO" else universe.get(m, "not in universe")
    )
    counts = (counts[["market", "country", "datapoints"]]
              .sort_values(["datapoints", "market"], ascending=[False, True])
              .reset_index(drop=True))

    if not verbose:
        return counts

    dim = proxy_dim(df["proxy_id"].iloc[0]) if len(df) else "?"
    years = f"{int(df['year'].min())}-{int(df['year'].max())}" if len(df) else "n/a"
    print(f"=== Coverage [{dim}] ===")
    print(f"Total datapoints : {len(df)}   |   years: {years}")

    # A GLO-only proxy is global by design; the 34-country check does not apply.
    if present == {"GLO"}:
        print("Global single series - country coverage not applicable")
        return counts

    print(f"Countries covered: {len(present & expected)} / {len(expected)}")
    if len(counts):
        print(f"Points per country: min {counts.datapoints.min()}, "
              f"max {counts.datapoints.max()}")

    missing = sorted(expected - present)
    if missing:
        print(f"\nMissing ({len(missing)}):")
        for iso in missing:
            print(f"   {iso}  {universe[iso]}")
    else:
        print("\nMissing: none - full coverage")

    extra = sorted(present - expected - {"GLO"})
    if extra:
        print(f"\nUnexpected markets: {extra}")
    return counts


def load(df: pd.DataFrame, workbook=WORKBOOK, sheet: str = SHEET) -> None:
    """Append a transformed proxy to the Historical Data sheet.

    Rows already present under the same (proxy_id, year) are skipped, so
    re-running a cell is safe and will not duplicate anything.
    """
    missing = [column for column in OUTPUT_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"Frame is missing contract columns: {missing}")
    if df.empty:
        print("  Nothing to write (empty frame)")
        return

    book = openpyxl.load_workbook(workbook)
    worksheet = book[sheet]

    edge = Side(style="thin", color="D0D0D0")
    border = Border(left=edge, right=edge, top=edge, bottom=edge)
    font = Font(name="Arial", size=10)
    centre = Alignment(horizontal="center", vertical="center")
    stripe = PatternFill("solid", start_color="F2F7FC")

    last_row = worksheet.max_row
    existing = set()
    written_rows = 0
    for row in range(2, last_row + 1):
        key = worksheet.cell(row=row, column=2).value
        if key:
            existing.add((key, worksheet.cell(row=row, column=4).value))
            written_rows += 1

    written = skipped = 0
    for _, record in df.iterrows():
        key = (record["proxy_id"], record["year"])
        if key in existing:
            skipped += 1
            continue

        last_row += 1
        written_rows += 1
        # `id` is the proxy dimension. Strip only the trailing market code so
        # sub-proxies (D7_1_USA -> D7_1) keep their numbering.
        values = [proxy_dim(record["proxy_id"])] + [record[c] for c in OUTPUT_COLUMNS]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=last_row, column=column, value=value)
            cell.font = font
            cell.alignment = centre
            cell.border = border
            if written_rows % 2 == 0:
                cell.fill = stripe
        written += 1

    book.save(workbook)
    print(f"  Written: {written} rows" + (f"   |   skipped: {skipped} already present"
                                          if skipped else ""))
